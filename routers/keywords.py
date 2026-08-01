"""
routers/keywords.py
===================
关键字管理路由（从 reply_server.py 迁移）。

路由清单：
- GET    /keywords/{cid}                    获取关键字列表（含 item_id）
- GET    /keywords-with-item-id/{cid}       获取含类型信息的关键词列表
- POST   /keywords/{cid}                    更新关键字（旧格式）
- POST   /keywords-with-item-id/{cid}       更新含商品 ID 的关键词（保留图片关键词）
- GET    /keywords-export/{cid}             导出关键词为 Excel
- POST   /keywords-import/{cid}             导入 Excel 关键词
- POST   /keywords/{cid}/image              添加图片关键词
- POST   /upload-image                      通用图片上传
- GET    /keywords-with-type/{cid}          获取含类型信息的关键词列表
- DELETE /keywords/{cid}/{index}            按索引删除关键词（图片关键词同时删文件）
- GET    /debug/keywords-table-info        调试：检查 keywords 表结构

设计要点：
- 权限：cid 必须属于当前用户（get_all_cookies 校验 或 get_cookie_details 校验）
- CookieManager：更新后需同步内存状态（cookie_manager.manager.update_keywords）
- 图片关键词：保存/删除时联动 image_manager 文件操作
- Excel 导入导出：pandas + openpyxl，空数据时生成带示例的模板
"""
import io
import time
import sqlite3
from typing import Dict, Any, Optional
from urllib.parse import quote

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from loguru import logger

from .deps import require_auth, server_error, client_error, log_with_user
from .models import KeywordIn, KeywordWithItemIdIn

router = APIRouter(tags=["keywords"])


def _db():
    from db_manager import db_manager
    return db_manager


def _mgr():
    """获取 CookieManager 单例；如未就绪抛 500"""
    import cookie_manager
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    return cookie_manager.manager


def _ensure_cookie_owned(cid: str, user_id: int) -> None:
    """校验 cid 属于当前用户，否则 403"""
    if cid not in _db().get_all_cookies(user_id):
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")


def _ensure_cookie_owned_v2(cid: str, user_id: int) -> None:
    """校验 cid 属于当前用户（通过 get_cookie_details），否则 404"""
    cookie_details = _db().get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != user_id:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")


# ------------------------- 查询 -------------------------

@router.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(require_auth)):
    """获取关键字列表（含 item_id）"""
    _mgr()  # 校验 CookieManager 就绪
    user_id = current_user['user_id']
    _ensure_cookie_owned(cid, user_id)

    item_keywords = _db().get_keywords_with_item_id(cid)
    return [
        {
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal",
        }
        for keyword, reply, item_id in item_keywords
    ]


@router.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(require_auth)):
    """获取含类型信息的关键词列表"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned(cid, user_id)

    keywords = _db().get_keywords_with_type(cid)
    return [
        {
            "keyword": kw['keyword'],
            "reply": kw['reply'],
            "item_id": kw['item_id'] or "",
            "type": kw['type'],
            "image_url": kw['image_url'],
        }
        for kw in keywords
    ]


@router.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(require_auth)):
    """获取含类型信息的关键词列表（原始格式）"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned_v2(cid, user_id)

    return _db().get_keywords_with_type(cid)


# ------------------------- 更新 -------------------------

@router.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(require_auth)):
    """更新关键字（旧格式，不含 item_id）"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned(cid, user_id)

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    _mgr().update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@router.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(require_auth)):
    """更新含商品 ID 的关键词列表（只保存文本关键词，保留图片关键词）"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned(cid, user_id)

    keywords_to_save = []
    keyword_set = set()
    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    try:
        success = _db().save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)

        if "已存在（图片关键词）" in error_msg:
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
            import re
            conflict_keyword = None
            conflict_type = None

            if "关键词唯一约束冲突" in error_msg:
                keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)
                if "通用关键词" in error_msg:
                    conflict_type = "通用关键词"
                elif "商品ID:" in error_msg:
                    item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

            if conflict_keyword and conflict_type:
                detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
            else:
                detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


# ------------------------- 删除 -------------------------

@router.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(require_auth)):
    """按索引删除关键词（图片关键词同时删除图片文件）"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned_v2(cid, user_id)

    try:
        keywords = _db().get_keywords_with_type(cid)
        if not (0 <= index < len(keywords)):
            raise HTTPException(status_code=400, detail="关键词索引无效")

        keyword_data = keywords[index]
        success = _db().delete_keyword_by_index(cid, index)
        if not success:
            raise HTTPException(status_code=400, detail="删除关键词失败")

        # 图片关键词 → 删除对应图片文件
        if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
            import image_manager
            image_manager.delete_image(keyword_data['image_url'])

        log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)
        return {"msg": "删除成功"}
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除关键词失败: {e}", current_user)
        raise server_error(e, "删除关键词")


# ------------------------- 图片关键词 -------------------------

@router.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(require_auth),
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    _mgr()
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")
    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    user_id = current_user['user_id']
    _ensure_cookie_owned_v2(cid, user_id)

    try:
        if not image.content_type or not image.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="请上传图片文件")

        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        import image_manager
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            raise HTTPException(status_code=400, detail="图片保存失败")
        logger.info(f"图片保存成功: {image_url}")

        # 检查重复
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if _db().check_keyword_duplicate(cid, keyword, normalized_item_id):
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存到数据库
        success = _db().save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)
        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise server_error(e, "添加图片关键词")


# ------------------------- 通用图片上传 -------------------------

@router.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(require_auth),
):
    """通用图片上传（用于卡券等功能）"""
    try:
        logger.info(f"接收到图片上传请求: filename={image.filename}")

        if not image.content_type or not image.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="请上传图片文件")

        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        import image_manager
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片上传成功: {image_url}")
        return {"message": "图片上传成功", "image_url": image_url}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图片上传失败: {e}")
        raise server_error(e, "图片上传")


# ------------------------- Excel 导入 / 导出 -------------------------

@router.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(require_auth)):
    """导出关键词为 Excel 文件"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned(cid, user_id)

    try:
        keywords = _db().get_keywords_with_type(cid)

        data = []
        for kw in keywords:
            if kw.get('type', 'text') == 'text':
                data.append({
                    '关键词': kw['keyword'],
                    '商品ID': kw['item_id'] or '',
                    '关键词内容': kw['reply'],
                })

        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            if data == []:
                worksheet = writer.sheets['关键词数据']
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'
                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'
                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"},
        )
    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise server_error(e, "导出关键词")


@router.post("/keywords-import/{cid}")
async def import_keywords(
    cid: str,
    file: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(require_auth),
):
    """导入 Excel 关键词到指定账号"""
    _mgr()
    user_id = current_user['user_id']
    _ensure_cookie_owned(cid, user_id)

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有文本关键词用于比较
        existing_keywords = _db().get_keywords_with_type(cid)
        existing_dict = {}
        for kw in existing_keywords:
            if kw.get('type', 'text') == 'text':
                key = f"{kw['keyword']}|{kw['item_id'] or ''}"
                existing_dict[key] = (kw['keyword'], kw['reply'], kw['item_id'])

        def clean_cell_value(value):
            if pd.isna(value):
                return ''
            if isinstance(value, float) and value == int(value):
                return str(int(value)).strip()
            return str(value).strip()

        import_data = []
        update_count = 0
        add_count = 0

        for _index, row in df.iterrows():
            keyword = clean_cell_value(row['关键词'])
            item_id = clean_cell_value(row['商品ID']) or None
            reply = clean_cell_value(row['关键词内容'])

            if not keyword:
                continue

            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                update_count += 1
            else:
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        success = _db().save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)
        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count,
        }
    except HTTPException:
        raise
    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise server_error(e, "导入关键词")


# ------------------------- 调试 -------------------------

@router.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(require_auth)):
    """调试：检查 keywords 表结构"""
    try:
        db = _db()
        conn = sqlite3.connect(db.db_path)
        cursor = conn.cursor()

        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns],
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise server_error(e, "检查表结构")
